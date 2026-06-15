"""
高考录取数据导入脚本（从xlsx文件）

数据来源：labolado/gaokao_2016-2020
格式：xlsx（每个省份一个文件，包含多个sheet：本科一批、本科二批、专科批等）
"""
import sqlite3
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("请安装openpyxl: pip install openpyxl")
    exit(1)

ROOT = Path(__file__).resolve().parents[1]
RAW_DATA_DIR = ROOT / "data" / "raw" / "gaokao_data"
DB_PATH = ROOT / "data" / "zx_advisor.db"

# 专业名称映射（xlsx中可能没有专业列，需要从sheet名称推断）
BATCH_MAPPING = {
    "普通类_本科一批": "本科一批",
    "普通类_本科二批": "本科二批",
    "普通类_专科批": "专科批",
    "普通类_本科提前批": "本科提前批",
}


def parse_score(value):
    """解析分数值"""
    if value is None or value == '--' or value == '':
        return None
    try:
        return int(float(str(value).replace(',', '')))
    except (ValueError, TypeError):
        return None


def parse_rank(value):
    """解析位次值"""
    if value is None or value == '--' or value == '':
        return None
    try:
        return int(float(str(value).replace(',', '')))
    except (ValueError, TypeError):
        return None


def get_university_id(conn, uni_name):
    """获取大学ID，如果不存在则创建"""
    cursor = conn.execute("SELECT id FROM universities WHERE name = ?", (uni_name,))
    row = cursor.fetchone()
    if row:
        return row[0]
    
    # 插入新大学
    cursor = conn.execute(
        "INSERT INTO universities (name, tier, city, tags) VALUES (?, ?, ?, ?)",
        (uni_name, "未知", "", "")
    )
    return cursor.lastrowid


def parse_xlsx_file(xlsx_path: Path, conn: sqlite3.Connection, dry_run: bool = False):
    """解析单个xlsx文件并导入数据"""
    # 从文件名提取省份和科类
    filename = xlsx_path.stem
    
    # 处理多种文件名格式
    # 格式1: "广东-理科.xlsx"
    # 格式2: "广东_理科.xlsx"（从目录名）
    match = re.match(r'(.+)[- _](.+)', filename)
    if not match:
        # 尝试从父目录名提取
        parent_name = xlsx_path.parent.name
        match = re.match(r'(.+)[_](.+)', parent_name)
        if not match:
            print(f"  ⚠️ 跳过无法解析的文件名: {filename}")
            return 0
    
    province = match.group(1)
    subject_type = match.group(2)
    
    print(f"\n  📊 处理: {province} - {subject_type}")
    
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    
    total_imported = 0
    
    # 遍历所有sheet（本科一批、本科二批等）
    for sheet_name in wb.sheetnames:
        batch = BATCH_MAPPING.get(sheet_name)
        if not batch:
            continue  # 跳过特殊批次（中外合作、三二分段等）
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        imported = 0
        for row in rows:
            if len(row) < 16:
                continue
            
            # 解析列
            # 0:省份, 1:城市, 2:985, 3:211, 4:双一流, 5:学校
            # 6:2020分数, 7:2020位次, 8:平均分
            # 9:2019分数, 10:2019位次, 11:平均分
            # 12:2018分数, 13:2018位次, 14:平均分
            # 15:2017分数, 16:2017位次, 17:平均分
            # 18:2016分数, 19:2016位次, 20:平均分
            
            uni_name = row[5] if len(row) > 5 else None
            if not uni_name or not isinstance(uni_name, str):
                continue
            
            # 获取大学ID
            uni_id = get_university_id(conn, uni_name)
            
            # 导入各年份数据
            years_data = [
                (2020, row[6] if len(row) > 6 else None, row[7] if len(row) > 7 else None),
                (2019, row[9] if len(row) > 9 else None, row[10] if len(row) > 10 else None),
                (2018, row[12] if len(row) > 12 else None, row[13] if len(row) > 13 else None),
                (2017, row[15] if len(row) > 15 else None, row[16] if len(row) > 16 else None),
                (2016, row[18] if len(row) > 18 else None, row[19] if len(row) > 19 else None),
            ]
            
            for year, score_val, rank_val in years_data:
                score = parse_score(score_val)
                rank = parse_rank(rank_val)
                
                if score is None:
                    continue
                
                # 插入数据（如果不存在）
                if not dry_run:
                    try:
                        conn.execute("""
                            INSERT OR IGNORE INTO admission_scores 
                            (university_id, province, subject_type, year, major_name, min_score, lowest_rank, data_source)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """, (uni_id, province, subject_type, year, "普通类", score, rank, "github_labolado"))
                        imported += 1
                    except Exception as e:
                        pass  # 忽略重复数据
        
        total_imported += imported
    
    wb.close()
    return total_imported


def import_all_xlsx(dry_run: bool = False):
    """导入所有xlsx文件"""
    print("=" * 60)
    print("📥 高考录取数据导入（xlsx → SQLite）")
    print("=" * 60)
    
    # 查找所有xlsx文件（处理目录结构）
    xlsx_files = []
    for item in RAW_DATA_DIR.rglob("*.xlsx"):
        if item.is_file():
            xlsx_files.append(item)
    
    if not xlsx_files:
        print("❌ 未找到xlsx文件，请先下载数据")
        return
    
    print(f"\n找到 {len(xlsx_files)} 个xlsx文件")
    
    # 连接数据库
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    
    total_imported = 0
    
    for xlsx_path in xlsx_files:
        try:
            imported = parse_xlsx_file(xlsx_path, conn, dry_run)
            total_imported += imported
        except Exception as e:
            print(f"  ❌ 处理失败: {xlsx_path.name} - {e}")
    
    if not dry_run:
        conn.commit()
        print(f"\n✅ 导入完成: {total_imported} 条记录")
        
        # 统计总数
        cursor = conn.execute("SELECT COUNT(*) FROM admission_scores")
        total = cursor.fetchone()[0]
        print(f"📊 数据库总记录数: {total}")
    else:
        print(f"\n🔍 预览模式: 将导入 {total_imported} 条记录")
    
    conn.close()


def show_stats():
    """显示数据统计"""
    if not DB_PATH.exists():
        print("数据库不存在")
        return
    
    conn = sqlite3.connect(str(DB_PATH))
    
    print("\n" + "=" * 60)
    print("📊 数据统计")
    print("=" * 60)
    
    # 按省份统计
    cursor = conn.execute("""
        SELECT province, COUNT(*) as count, 
               MIN(year) as min_year, MAX(year) as max_year
        FROM admission_scores 
        GROUP BY province 
        ORDER BY count DESC
    """)
    
    print("\n按省份统计:")
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1]}条 ({row[2]}-{row[3]})")
    
    # 按年份统计
    cursor = conn.execute("""
        SELECT year, COUNT(*) as count
        FROM admission_scores 
        GROUP BY year 
        ORDER BY year DESC
    """)
    
    print("\n按年份统计:")
    for row in cursor.fetchall():
        print(f"  {row[0]}年: {row[1]}条")
    
    conn.close()


if __name__ == "__main__":
    import sys
    
    if "--stats" in sys.argv:
        show_stats()
    elif "--dry-run" in sys.argv:
        import_all_xlsx(dry_run=True)
    else:
        import_all_xlsx(dry_run=False)
